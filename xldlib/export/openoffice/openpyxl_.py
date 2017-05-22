'''
    Export/OpenOffice/openpyxl_
    ___________________________

    OpenPyXl engine for the Open Office writer.

    :copyright: (c) 2015 The Regents of the University of California.
    :license: GNU GPL, see licenses/GNU GPLv3.txt for more details.
'''

# load modules/submodules
import openpyxl

from xldlib.utils import logger

from . import base


# FORMATS
# -------

HEADER = {
    'font': openpyxl.styles.Font(bold=True, underline='single'),
    'alignment': openpyxl.styles.Alignment(horizontal='center')
}


# HELPERS
# -------

def num_to_col(div):
    '''Processes an integer to an excel column'''

    col = ""
    while div > 0:
        # grab the remainder and adjust to string
        module = (div-1) % 26
        col = chr(ord('A')+module)+col
        div = int((div-module)/26)
    return col


# OBJECTS
# -------


@logger.init('spreadsheet', 'DEBUG')
class Worksheet(base.Worksheet):
    '''Standardized worksheet object for OpenPyXl'''

    def __init__(self, worksheet, workbook):
        super(Worksheet, self).__init__(worksheet, workbook)

        self.column_dimensions = worksheet.column_dimensions

        self.set_version()

    #  INITIALIZERS

    @classmethod
    @logger.call('spreadsheet', 'debug')
    def new(cls, workbook, index=None, title='Sheet'):
        worksheet = workbook.create_sheet(index, title)
        return cls(worksheet, workbook)

    #    SETTERS

    def set_version(self):
        '''Sets the current OpenPyXl version and the index offsets'''

        self.version = openpyxl.__version__.split('.')[0]
        # OpenPyXl 2.x uses 1-indexing, 1.x uses 0-indexing
        self.index_offset = self.version == '2'

    def set_styles(self, dataframe, columns):
        '''Binds styles to the class for quick lookups'''

        import pdb, sys; pdb.Pdb(stdout=sys.__stdout__).set_trace()
        self.number_formatting = {}
        for index, key in enumerate(dataframe):
            data = columns.get(key)
            if data is not None and 'num_format' in data.formatting:
                self.number_formatting[index] = data.formatting['num_format']

    def set_columnsizes(self, dataframe, columns):
        '''Sets the sizes for each column in the sheet'''

        # TODO: TEST
        for index, key in enumerate(dataframe):
            data = columns.get(key)
            if data is not None:
                column = num_to_col(index + 1)

                try:
                    self.worksheet.column_dimensions[column].width = data.size
                except KeyError:
                    self.worksheet.get_cell(0, 0)
                    self.worksheet.column_dimensions[column].width = data.size

    def set_cellstyles(self, cell, styles):
        '''Writes the OpenPyXl styles to cell prior to adding data'''

        for key, value in styles.items():
            if self.version == '1':
                # everything is stored in cell.style for 1.x - 1.8x
                setattr(cell.style, key, value)
            elif self.version == '2':
                # everything is stored in cell.style for > 2.x
                setattr(cell, key, value)

    def set_cellformats(self, cell, column):
        '''Writes the OpenPyXl number formats to cell prior to adding data'''

        formatting = self.number_formatting.get(column)
        if formatting is not None:
            if self.version == '1':
                cell.style.number_format.format_code = formatting
            elif self.version == '2':
                cell.number_format = formatting

    def set_padding(self):
        '''Adds padding to the spreadsheet to avoid column truncation'''

        self.get_cell(0, 1000)

    #    GETTERS

    def get_cell(self, row, column):
        '''Returns the cell from the current worksheet'''

        return self.worksheet.cell(row=row+self.index_offset,
            column=column+self.index_offset)

    #    WRITERS

    def _write_cell(self, row, column, value, header):
        '''Writes a value to an OpenPyXl cell'''

        cell = self.get_cell(row, column)
        if header:
            import pdb, sys; pdb.Pdb(stdout=sys.__stdout__).set_trace()
            self.set_cellstyles(cell, HEADER)
        else:
            import pdb, sys; pdb.Pdb(stdout=sys.__stdout__).set_trace()
            self.set_cellformats(cell, column)

        cell.value = value

    #    HELPERS

    def merge(self, first_row, first_column, last_row, last_column, data):
        '''Provides a merge range for the worksheet'''

        if data and data != ' ':
            self._write_cell(first_row, first_column, data, header=True)
            self.worksheet.merge_cells(start_row=first_row+self.index_offset,
                start_col=first_column+self.index_offset,
                end_row=last_row+self.index_offset,
                end_col=last_column+self.index_offset)


@logger.init('spreadsheet', 'DEBUG')
class Workbook(base.Workbook):
    '''Workbook object for the Open Office writer'''

    def __init__(self, path):
        super(Workbook, self).__init__(path)

        self.workbook = openpyxl.Workbook()
        self.__removesheet()

    #     MAGIC

    def __getitem__(self, index):
        '''Workbook[index] -> Worksheet'''

        worksheet = self.workbook.worksheets[index]
        return Worksheet(worksheet, self)

    #  PUBLIC FUNCTIONS

    def add_worksheet(self, index, title, dataframe=None):
        '''Adds a worksheet to the workbook at the given index'''

        title = self._titlechecker(title)
        worksheet = Worksheet.new(self, index, title)
        if dataframe is not None:
            worksheet.setdataframe(dataframe)

    def create_sheet(self, index, title):
        self.workbook.create_sheet(index, title=title)

    def save(self):
        self.workbook.save(self.path)

    #    HELPERS

    def __removesheet(self):
        '''Removes the sheet created on workbook initialization'''

        sheet = self.workbook.get_sheet_by_name("Sheet")
        if sheet is not None:
            self.workbook.remove_sheet(sheet)
